"""
exporter.py  —  Financial-Grade Formatting
───────────────────────────────────────────
Generates one Excel workbook per company with 6 sheets:

  Formatted  │ Income Statement  │ Balance Sheet  │ Cash Flow
  Raw        │ raw_Income Statement │ raw_Balance Sheet │ raw_Cash Flow

Design principles:
  · No gridlines — clean white canvas
  · Accounting number format: negatives in (parentheses), zeros as "—"
  · Borders only on subtotal / total rows — no cell-level box-drawing
  · Muted professional color palette (deep navy headers, ice-blue fills)
  · Calibri throughout; bold for subtotals / totals; 12 pt for banner
  · Missing values shown as "—" (gray italic) — no yellow fill
  · Print-ready: landscape, fit-to-width, header/footer suppressed
"""

import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import (
    get_company_info, DEFAULT_YEARS, OUTPUT_DIR,
    SHEET_NAMES, RAW_SHEET_NAMES, EN_RAW_SHEET_NAMES,
)

# ── Number format ─────────────────────────────────────────────────────────────
# Positive: 1,234   Negative: (1,234)   Zero: —
FIN_NUMBER_FORMAT = '#,##0_);(#,##0);"-"'

# ── Professional color palette ────────────────────────────────────────────────
C_BANNER    = "1F3864"   # Deep navy    — title banner
C_COL_HDR   = "2E4057"   # Slate blue   — column header row
C_SECTION   = "EEF2F8"   # Ghost white  — section headers (ASSETS, etc.)
C_SUBTOTAL  = "DBE8F5"   # Ice blue     — subtotal rows
C_TOTAL     = "C5D9F1"   # Sky blue     — grand total rows
C_WHITE     = "FFFFFF"
C_TEXT_NAVY = "1F3864"   # Deep navy    — section header label text
C_TEXT_GRAY = "9CA3AF"   # Medium gray  — "—" missing value text
C_BORDER    = "B8C9DB"   # Cool gray    — border lines

# Raw-sheet accent (dark slate — distinct from company color)
C_RAW_HDR   = "3D4B5C"

# English raw-sheet accent (dark forest green)
C_EN_HDR    = "2D6A4F"


# ── Vietnamese → English translation table ────────────────────────────────────
# Keys are stripped/normalized Vietnamese labels as they appear in raw_lines.
# Values are plain English equivalents suitable for a financial audience.

VI_TO_EN: dict[str, str] = {
    # ── UPPERCASE / ALL-CAPS SECTION HEADERS (some companies use these) ─────────
    "TÀI SẢN NGẮN HẠN":                                              "Current Assets",
    "TÀI SẢN DÀI HẠN":                                               "Non-current Assets",
    "TỔNG TÀI SẢN":                                                   "Total Assets",
    "NỢ PHẢI TRẢ":                                                    "Total Liabilities",
    "NỢ NGẮN HẠN":                                                    "Current Liabilities",
    "NỢ DÀI HẠN":                                                     "Non-current Liabilities",
    "VỐN CHỦ SỞ HỮU":                                                "Equity",
    "LƯU CHUYỂN TIỀN TỪ HOẠT ĐỘNG KINH DOANH":                       "Cash Flows from Operating Activities",
    "LƯU CHUYỂN TIỀN TỪ HOẠT ĐỘNG ĐẦU TƯ":                          "Cash Flows from Investing Activities",
    "LƯU CHUYỂN TIỀN TỪ HOẠT ĐỘNG TÀI CHÍNH":                        "Cash Flows from Financing Activities",
    "LƯU CHUYỂN TIỀN THUẦN TRONG KỲ":                                 "Net Increase / (Decrease) in Cash",
    "TIỀN VÀ CÁC KHOẢN TƯƠNG ĐƯƠNG TIỀN":                            "Cash and Cash Equivalents",
    "HÀNG TỒN KHO":                                                   "Inventories",
    "TÀI SẢN CỐ ĐỊNH":                                                "Fixed Assets",
    "BẤT ĐỘNG SẢN ĐẦU TƯ":                                           "Investment Properties",
    "ĐẦU TƯ TÀI CHÍNH DÀI HẠN":                                      "Long-term Financial Investments",
    "ĐẦU TƯ TÀI CHÍNH NGẮN HẠN":                                     "Short-term Financial Investments",

    # ── VARIANT SPELLINGS (different companies phrase these differently) ──────
    "Giá trị khấu hao lũy kế":                                        "Less: Accumulated Depreciation",
    "Giá trị hao mòn lũy kế":                                         "Less: Accumulated Depreciation",
    "Hao mòn lũy kế":                                                 "Less: Accumulated Depreciation",
    "Khấu hao lũy kế":                                                "Less: Accumulated Depreciation",
    "Giá trị còn lại":                                                "Net Book Value",
    "Các khoản giảm trừ doanh thu":                                   "Less: Revenue Deductions",
    "Giảm trừ doanh thu":                                             "Less: Revenue Deductions",
    "Chiết khấu thương mại":                                          "Trade Discounts",
    "Hàng bán bị trả lại":                                            "Sales Returns",
    "Giảm giá hàng bán":                                              "Sales Allowances",
    "Lợi nhuận từ hoạt động kinh doanh trước thay đổi vốn lưu động": "Operating Profit Before Working Capital Changes",
    "Lưu chuyển tiền trước thay đổi vốn lưu động":                   "Cash Before Working Capital Changes",
    "Tăng (+), giảm (-) các khoản phải thu":                         "Change in Receivables",
    "Tăng (+), giảm (-) hàng tồn kho":                               "Change in Inventories",
    "Tăng (+), giảm (-) các khoản phải trả":                         "Change in Payables",
    "Tăng (+), giảm (-) chi phí trả trước":                          "Change in Prepaid Expenses",
    "Các khoản điều chỉnh khác":                                      "Other Adjustments",
    "Tiền lãi cho vay và cổ tức đã thu":                              "Cash Received — Interest & Dividends",
    "Tiền thu lãi cho vay và cổ tức":                                 "Cash Received — Interest & Dividends",
    "Tiền chi mua sắm TSCĐ":                                         "Cash Paid — Fixed Asset Purchases",
    "Tiền thu từ bán TSCĐ":                                           "Cash Received from Asset Sales",
    "Tiền chi đầu tư vào các đơn vị khác":                           "Cash Paid — Investments in Other Entities",
    "Tiền thu hồi đầu tư vào các đơn vị khác":                       "Cash Received from Investment Returns",
    "Tiền thu từ phát hành cổ phiếu":                                 "Cash Received from Share Issuance",
    "Tiền chi mua lại cổ phiếu quỹ":                                  "Cash Paid for Treasury Shares",
    "Vốn góp của chủ sở hữu":                                         "Contributed Capital",
    "Vốn đầu tư của chủ sở hữu":                                      "Owner's Invested Capital",
    "Thặng dư vốn cổ phần (vốn khác)":                               "Share Premium (Other Capital)",
    "Chênh lệch đánh giá lại tài sản":                                "Asset Revaluation Surplus",
    "Quỹ dự phòng tài chính":                                         "Financial Reserve Fund",
    "Quỹ dự trữ bắt buộc":                                            "Statutory Reserve Fund",
    "Quỹ khác thuộc vốn chủ sở hữu":                                  "Other Equity Reserves",
    "Phần lãi (lỗ) trong liên doanh, liên kết":                       "Share of Profit/(Loss) from JVs & Associates",
    "Lãi/(lỗ) từ công ty liên kết":                                   "Share of Profit/(Loss) from Associates",
    "Dự phòng rủi ro":                                                "Risk Provision",
    "Dự phòng giảm giá đầu tư":                                       "Provision for Decline in Investments",
    "Dự phòng nợ phải thu khó đòi":                                   "Provision for Doubtful Debts",
    "Chi phí khấu hao":                                               "Depreciation Expense",
    "Khấu hao TSCĐ":                                                  "Depreciation of Fixed Assets",
    "Phân bổ công cụ dụng cụ":                                        "Amortization of Tools & Equipment",
    "Chi phí phân bổ":                                                "Amortization Expense",
    "Tài sản ngắn hạn khác (*)":                                      "Other Current Assets",
    "Tài sản dài hạn khác (*)":                                       "Other Non-current Assets",
    "Phải thu khác":                                                   "Other Receivables",
    "Phải trả khác":                                                   "Other Payables",
    "Tổng tài sản":                                                    "Total Assets",
    "Tổng nguồn vốn":                                                  "Total Liabilities & Equity",
    "Tổng cộng tài sản":                                               "Total Assets",
    "Tổng cộng nguồn vốn":                                             "Total Liabilities & Equity",

    # ── INCOME STATEMENT ────────────────────────────────────────────────────────
    "Doanh thu bán hàng và cung cấp dịch vụ":                        "Revenue from Goods & Services",
    "Doanh thu thuần về bán hàng và cung cấp dịch vụ":               "Net Revenue",
    "Doanh thu thuần":                                                "Net Revenue",
    "Doanh thu":                                                      "Revenue",
    "Giá vốn hàng bán":                                               "Cost of Goods Sold",
    "Lợi nhuận gộp về bán hàng và cung cấp dịch vụ":                 "Gross Profit",
    "Lợi nhuận gộp":                                                  "Gross Profit",
    "Doanh thu hoạt động tài chính":                                  "Financial Income",
    "Doanh thu tài chính":                                            "Financial Income",
    "Chi phí tài chính":                                              "Financial Expenses",
    "Trong đó: Chi phí lãi vay":                                      "  of which: Interest Expense",
    "Chi phí lãi vay":                                                "Interest Expense",
    "Chi phí bán hàng":                                               "Selling Expenses",
    "Chi phí quản lý doanh nghiệp":                                   "General & Administrative Expenses",
    "Chi phí quản lý":                                                "Management Expenses",
    "Lợi nhuận thuần từ hoạt động kinh doanh":                        "Operating Profit",
    "Lợi nhuận từ hoạt động kinh doanh":                              "Operating Profit",
    "Thu nhập khác":                                                  "Other Income",
    "Chi phí khác":                                                   "Other Expenses",
    "Lợi nhuận khác":                                                 "Other Profit",
    "Phần lãi (lỗ) trong công ty liên kết, liên doanh":               "Share of Profit from Associates & JVs",
    "Phần lãi trong công ty liên kết, liên doanh":                    "Share of Profit from Associates & JVs",
    "Lợi nhuận trước thuế":                                           "Profit Before Tax",
    "Chi phí thuế thu nhập doanh nghiệp hiện hành":                   "Current Corporate Income Tax",
    "Chi phí thuế thu nhập doanh nghiệp hoãn lại":                    "Deferred Corporate Income Tax",
    "Chi phí thuế thu nhập doanh nghiệp":                             "Corporate Income Tax",
    "Lợi nhuận sau thuế thu nhập doanh nghiệp":                       "Net Profit After Tax",
    "Lợi nhuận sau thuế":                                             "Net Profit After Tax",
    "Lợi nhuận sau thuế của cổ đông của công ty mẹ":                  "Net Profit — Parent Company Shareholders",
    "Lợi nhuận sau thuế của cổ đông công ty mẹ":                      "Net Profit — Parent Company Shareholders",
    "Lợi ích của cổ đông không kiểm soát":                            "Net Profit — Non-controlling Interest",
    "Lợi ích cổ đông không kiểm soát":                                "Net Profit — Non-controlling Interest",
    "Lãi cơ bản trên cổ phiếu":                                       "Basic Earnings Per Share (VND)",
    "Lãi suy giảm trên cổ phiếu":                                     "Diluted Earnings Per Share (VND)",

    # ── BALANCE SHEET — ASSETS ──────────────────────────────────────────────────
    "TÀI SẢN":                                                        "ASSETS",
    "TỔNG CỘNG TÀI SẢN":                                              "TOTAL ASSETS",
    "Tài sản ngắn hạn":                                               "Current Assets",
    "Tổng tài sản ngắn hạn":                                          "Total Current Assets",
    "Tiền và các khoản tương đương tiền":                             "Cash and Cash Equivalents",
    "Tiền":                                                           "Cash",
    "Các khoản tương đương tiền":                                     "Cash Equivalents",
    "Đầu tư tài chính ngắn hạn":                                      "Short-term Financial Investments",
    "Chứng khoán kinh doanh":                                         "Trading Securities",
    "Dự phòng giảm giá chứng khoán kinh doanh":                       "Provision for Decline in Trading Securities",
    "Đầu tư nắm giữ đến ngày đáo hạn":                               "Held-to-Maturity Investments",
    "Các khoản phải thu ngắn hạn":                                    "Short-term Receivables",
    "Phải thu ngắn hạn của khách hàng":                               "Trade Receivables (Short-term)",
    "Phải thu khách hàng":                                            "Trade Receivables",
    "Trả trước cho người bán ngắn hạn":                               "Prepayments to Suppliers (Short-term)",
    "Trả trước cho người bán":                                        "Prepayments to Suppliers",
    "Phải thu nội bộ ngắn hạn":                                       "Inter-company Receivables (Short-term)",
    "Phải thu về cho vay ngắn hạn":                                   "Loan Receivables (Short-term)",
    "Phải thu ngắn hạn khác":                                         "Other Short-term Receivables",
    "Dự phòng phải thu ngắn hạn khó đòi":                            "Provision for Short-term Doubtful Debts",
    "Tài sản thiếu chờ xử lý":                                        "Assets Awaiting Resolution",
    "Hàng tồn kho":                                                   "Inventories",
    "Dự phòng giảm giá hàng tồn kho":                                 "Provision for Inventory Write-down",
    "Tài sản ngắn hạn khác":                                          "Other Current Assets",
    "Chi phí trả trước ngắn hạn":                                     "Prepaid Expenses (Short-term)",
    "Thuế GTGT được khấu trừ":                                        "VAT Recoverable",
    "Thuế và các khoản khác phải thu Nhà nước":                       "Taxes Receivable from State",
    "Giao dịch mua bán lại trái phiếu Chính phủ":                     "Repo / Reverse Repo — Government Bonds",
    "Tài sản dài hạn":                                                "Non-current Assets",
    "Tổng tài sản dài hạn":                                           "Total Non-current Assets",
    "Các khoản phải thu dài hạn":                                     "Long-term Receivables",
    "Phải thu dài hạn của khách hàng":                                "Long-term Trade Receivables",
    "Vốn kinh doanh ở đơn vị trực thuộc":                             "Capital at Subsidiary",
    "Phải thu dài hạn nội bộ":                                        "Long-term Inter-company Receivables",
    "Phải thu về cho vay dài hạn":                                    "Long-term Loan Receivables",
    "Phải thu dài hạn khác":                                          "Other Long-term Receivables",
    "Dự phòng phải thu dài hạn khó đòi":                             "Provision for Long-term Doubtful Debts",
    "Tài sản cố định":                                                "Fixed Assets",
    "Tài sản cố định hữu hình":                                       "Tangible Fixed Assets",
    "Tài sản cố định thuê tài chính":                                 "Finance Lease Assets",
    "Tài sản cố định vô hình":                                        "Intangible Fixed Assets",
    "Nguyên giá":                                                     "Cost (Gross)",
    "Giá trị hao mòn lũy kế":                                         "Less: Accumulated Depreciation",
    "Bất động sản đầu tư":                                            "Investment Properties",
    "Tài sản dở dang dài hạn":                                        "Long-term Work-in-Progress",
    "Chi phí sản xuất, kinh doanh dở dang dài hạn":                   "Long-term Work-in-Progress",
    "Chi phí xây dựng cơ bản dở dang":                               "Construction in Progress",
    "Đầu tư tài chính dài hạn":                                       "Long-term Financial Investments",
    "Đầu tư vào công ty con":                                         "Investment in Subsidiaries",
    "Đầu tư vào công ty liên kết, liên doanh":                        "Investment in Associates & JVs",
    "Đầu tư góp vốn vào đơn vị khác":                                "Other Equity Investments",
    "Dự phòng đầu tư tài chính dài hạn":                             "Provision for Long-term Financial Investments",
    "Tài sản dài hạn khác":                                           "Other Non-current Assets",
    "Chi phí trả trước dài hạn":                                      "Long-term Prepaid Expenses",
    "Tài sản thuế thu nhập hoãn lại":                                 "Deferred Tax Assets",
    "Thiết bị, vật tư, phụ tùng thay thế dài hạn":                   "Long-term Spare Parts & Materials",
    "Lợi thế thương mại":                                             "Goodwill",

    # ── BALANCE SHEET — LIABILITIES & EQUITY ────────────────────────────────────
    "NGUỒN VỐN":                                                      "LIABILITIES & EQUITY",
    "TỔNG CỘNG NGUỒN VỐN":                                            "TOTAL LIABILITIES & EQUITY",
    "Nợ phải trả":                                                    "Total Liabilities",
    "Tổng nợ phải trả":                                               "Total Liabilities",
    "Nợ ngắn hạn":                                                    "Current Liabilities",
    "Tổng nợ ngắn hạn":                                               "Total Current Liabilities",
    "Phải trả người bán ngắn hạn":                                    "Trade Payables (Short-term)",
    "Phải trả người bán":                                             "Trade Payables",
    "Người mua trả tiền trước ngắn hạn":                              "Advances from Customers (Short-term)",
    "Người mua trả tiền trước":                                       "Advances from Customers",
    "Thuế và các khoản phải nộp Nhà nước":                            "Taxes Payable to State",
    "Phải trả người lao động":                                        "Payables to Employees",
    "Chi phí phải trả ngắn hạn":                                      "Accrued Expenses (Short-term)",
    "Chi phí phải trả":                                               "Accrued Expenses",
    "Phải trả nội bộ ngắn hạn":                                       "Inter-company Payables (Short-term)",
    "Phải trả theo tiến độ kế hoạch hợp đồng xây dựng":              "Payables per Construction Contract",
    "Doanh thu chưa thực hiện ngắn hạn":                              "Deferred Revenue (Short-term)",
    "Phải trả ngắn hạn khác":                                         "Other Short-term Payables",
    "Vay và nợ thuê tài chính ngắn hạn":                              "Short-term Borrowings & Finance Leases",
    "Dự phòng phải trả ngắn hạn":                                     "Short-term Provisions",
    "Quỹ khen thưởng, phúc lợi":                                      "Bonus and Welfare Fund",
    "Quỹ bình ổn giá":                                                "Price Stabilization Fund",
    "Nợ dài hạn":                                                     "Non-current Liabilities",
    "Tổng nợ dài hạn":                                                "Total Non-current Liabilities",
    "Phải trả người bán dài hạn":                                     "Long-term Trade Payables",
    "Người mua trả tiền trước dài hạn":                               "Long-term Advances from Customers",
    "Chi phí phải trả dài hạn":                                       "Long-term Accrued Expenses",
    "Phải trả nội bộ về vốn kinh doanh":                              "Inter-company Payables — Capital",
    "Phải trả dài hạn nội bộ":                                        "Long-term Inter-company Payables",
    "Doanh thu chưa thực hiện dài hạn":                               "Long-term Deferred Revenue",
    "Phải trả dài hạn khác":                                          "Other Long-term Payables",
    "Vay và nợ thuê tài chính dài hạn":                               "Long-term Borrowings & Finance Leases",
    "Trái phiếu chuyển đổi":                                          "Convertible Bonds",
    "Cổ phiếu ưu đãi":                                                "Preference Shares",
    "Thuế thu nhập hoãn lại phải trả":                                "Deferred Tax Liabilities",
    "Dự phòng phải trả dài hạn":                                      "Long-term Provisions",
    "Quỹ phát triển khoa học và công nghệ":                           "Science & Technology Development Fund",
    "Vốn chủ sở hữu":                                                 "Equity",
    "Tổng vốn chủ sở hữu":                                            "Total Equity",
    "Vốn chủ sở hữu của cổ đông công ty mẹ":                         "Equity of Parent Company Shareholders",
    "Vốn của chủ sở hữu":                                             "Owner's Equity",
    "Vốn góp của chủ sở hữu":                                         "Contributed Capital",
    "Vốn điều lệ":                                                    "Charter Capital",
    "Thặng dư vốn cổ phần":                                           "Share Premium",
    "Vốn khác của chủ sở hữu":                                        "Other Owner's Capital",
    "Cổ phiếu quỹ":                                                   "Treasury Shares",
    "Chênh lệch đánh giá lại tài sản":                                "Asset Revaluation Surplus",
    "Chênh lệch tỷ giá hối đoái":                                     "Foreign Exchange Difference",
    "Quỹ đầu tư phát triển":                                          "Development Investment Fund",
    "Quỹ hỗ trợ sắp xếp doanh nghiệp":                               "Enterprise Restructuring Fund",
    "Lợi nhuận sau thuế chưa phân phối":                              "Retained Earnings",
    "Lợi nhuận chưa phân phối":                                       "Retained Earnings",
    "Lợi nhuận sau thuế chưa phân phối lũy kế đến cuối kỳ trước":    "Retained Earnings — Prior Periods",
    "Lợi nhuận sau thuế của kỳ này":                                  "Net Profit — Current Period",
    "Nguồn vốn đầu tư xây dựng cơ bản":                              "Capital for Construction Investment",
    "Lợi ích cổ đông không kiểm soát":                                "Non-controlling Interest",

    # ── CASH FLOW ───────────────────────────────────────────────────────────────
    "Lưu chuyển tiền từ hoạt động kinh doanh":                        "Cash Flows from Operating Activities",
    "Lợi nhuận trước thuế thu nhập doanh nghiệp":                     "Profit Before Tax",
    "Điều chỉnh cho các khoản":                                       "Adjustments for:",
    "Khấu hao TSCĐ và bất động sản đầu tư":                          "Depreciation — Fixed Assets & Investment Properties",
    "Khấu hao tài sản cố định":                                       "Depreciation of Fixed Assets",
    "Khấu hao TSCĐ":                                                  "Depreciation of Fixed Assets",
    "Phân bổ lợi thế thương mại":                                     "Amortization of Goodwill",
    "Dự phòng":                                                       "Provisions",
    "Lãi, lỗ chênh lệch tỷ giá hối đoái chưa thực hiện":             "Unrealized FX Gains / (Losses)",
    "Lãi, lỗ từ hoạt động đầu tư":                                    "Gains / (Losses) from Investing Activities",
    "Lãi tiền gửi, tiền cho vay":                                     "Interest Income — Deposits & Loans",
    "Lợi nhuận từ hoạt động kinh doanh trước thay đổi vốn lưu động":  "Operating Profit Before Working Capital Changes",
    "Tăng, giảm các khoản phải thu":                                  "Change in Receivables",
    "Tăng, giảm hàng tồn kho":                                        "Change in Inventories",
    "Tăng, giảm các khoản phải trả (không kể lãi vay và thuế TNDN)": "Change in Payables (excl. interest & tax)",
    "Tăng, giảm các khoản phải trả":                                  "Change in Payables",
    "Tăng, giảm chi phí trả trước":                                   "Change in Prepaid Expenses",
    "Tăng, giảm chứng khoán kinh doanh":                              "Change in Trading Securities",
    "Tiền lãi vay đã trả":                                            "Interest Paid",
    "Thuế thu nhập doanh nghiệp đã nộp":                              "Corporate Income Tax Paid",
    "Tiền thu khác từ hoạt động kinh doanh":                          "Other Cash Received from Operations",
    "Tiền chi khác cho hoạt động kinh doanh":                         "Other Cash Paid for Operations",
    "Lưu chuyển tiền thuần từ hoạt động kinh doanh":                  "Net Cash Flows from Operating Activities",
    "Lưu chuyển tiền từ hoạt động đầu tư":                            "Cash Flows from Investing Activities",
    "Tiền chi để mua sắm, xây dựng TSCĐ và các tài sản dài hạn khác": "Cash Paid — Fixed Assets & Long-term Assets",
    "Tiền thu từ thanh lý, nhượng bán TSCĐ và các tài sản dài hạn khác": "Cash Received from Disposal of Fixed Assets",
    "Tiền chi cho vay và mua các công cụ nợ của đơn vị khác":         "Cash Paid — Loans & Debt Instruments",
    "Tiền thu hồi cho vay, bán lại các công cụ nợ của đơn vị khác":   "Cash Received from Loan Repayments",
    "Tiền chi đầu tư góp vốn vào đơn vị khác":                       "Cash Paid — Equity Investments",
    "Tiền thu hồi đầu tư góp vốn vào đơn vị khác":                   "Cash Received from Equity Investment Returns",
    "Tiền thu lãi cho vay, cổ tức và lợi nhuận được chia":            "Cash Received — Interest, Dividends & Profit Shares",
    "Lưu chuyển tiền thuần từ hoạt động đầu tư":                      "Net Cash Flows from Investing Activities",
    "Lưu chuyển tiền từ hoạt động tài chính":                         "Cash Flows from Financing Activities",
    "Tiền thu từ phát hành cổ phiếu, nhận vốn góp của chủ sở hữu":   "Cash Received — Share Issuance / Owner Capital",
    "Tiền chi trả vốn góp cho các chủ sở hữu, mua lại cổ phiếu của doanh nghiệp đã phát hành": "Cash Paid — Capital Repayment / Share Buyback",
    "Tiền vay ngắn hạn, dài hạn nhận được":                           "Cash Received from Borrowings",
    "Tiền thu từ đi vay":                                             "Cash Received from Borrowings",
    "Tiền chi trả nợ gốc vay":                                        "Cash Paid — Loan Principal Repayment",
    "Tiền trả nợ gốc vay":                                            "Cash Paid — Loan Principal Repayment",
    "Tiền chi trả nợ thuê tài chính":                                 "Cash Paid — Finance Lease Liabilities",
    "Cổ tức, lợi nhuận đã trả cho chủ sở hữu":                       "Dividends & Profit Paid to Owners",
    "Lưu chuyển tiền thuần từ hoạt động tài chính":                   "Net Cash Flows from Financing Activities",
    "Lưu chuyển tiền thuần trong kỳ":                                 "Net Increase / (Decrease) in Cash",
    "Tiền và tương đương tiền đầu kỳ":                                "Cash & Equivalents — Beginning of Period",
    "Ảnh hưởng của thay đổi tỷ giá hối đoái quy đổi ngoại tệ":       "Effect of Exchange Rate Changes on Cash",
    "Tiền và tương đương tiền cuối kỳ":                               "Cash & Equivalents — End of Period",
}


# Matches leading prefixes that companies prepend to line item labels, e.g.:
#   "A. TÀI SẢN NGẮN HẠN"  →  "TÀI SẢN NGẮN HẠN"
#   "I. Tiền và các khoản…" →  "Tiền và các khoản…"
#   "1. Tiền"               →  "Tiền"
#   "II. "  "III. "  "2.3. " etc.
_PREFIX_RE = re.compile(
    r"^(?:[A-Z]{1,3}\.(?:[A-Z]{1,3}\.)*|[IVXLC]+\.|[0-9]+(?:\.[0-9]+)*\.)\s+"
)


def _strip_prefix(label: str) -> str:
    """Remove a leading alphanumeric/Roman-numeral section prefix if present."""
    return _PREFIX_RE.sub("", label).strip()


def _translate_vi(label: str) -> str:
    """
    Translate a Vietnamese financial statement label to English.

    Lookup order:
    1. Exact match
    2. Stripped whitespace match
    3. After removing leading section prefix (e.g. "I. ", "A. ", "1. ")
    4. Case-insensitive match on the de-prefixed core
    5. Unfound → return original prefixed with ⚑ so reviewers can spot gaps
    """
    if not label:
        return ""

    # 1. Exact
    if label in VI_TO_EN:
        return VI_TO_EN[label]

    # 2. Stripped whitespace
    stripped = label.strip()
    if stripped in VI_TO_EN:
        return VI_TO_EN[stripped]

    # 3. Strip leading section prefix then exact lookup
    core = _strip_prefix(stripped)
    if core and core in VI_TO_EN:
        return VI_TO_EN[core]

    # 4. Case-insensitive on core
    lower_core = core.lower()
    for vi, en in VI_TO_EN.items():
        if vi.lower() == lower_core:
            return en

    # 5. Not found — keep original so no data is hidden
    return f"⚑ {stripped}"


# ── Canonical row definitions ─────────────────────────────────────────────────
# (display_label, field_key, row_type)
# row_type: 'header' | 'item' | 'subtotal' | 'total'
# field_key = None for section-header rows (no numeric data)

IS_ROWS = [
    ("Net Revenue",                   "net_revenue",                   "subtotal"),
    ("Cost of Goods Sold",            "cost_of_goods_sold",            "item"),
    ("Gross Profit",                  "gross_profit",                  "subtotal"),
    ("Selling Expenses",              "selling_expenses",              "item"),
    ("General & Admin Expenses",      "general_admin_expenses",        "item"),
    ("Operating Profit (EBIT)",       "operating_profit",              "subtotal"),
    ("Financial Income",              "financial_income",              "item"),
    ("Financial Expenses",            "financial_expenses",            "item"),
    ("  of which: Interest Expense",  "interest_expense",              "item"),
    ("Profit from Associates",        "profit_from_associates",        "item"),
    ("Other Income",                  "other_income",                  "item"),
    ("Other Expenses",                "other_expenses",                "item"),
    ("Profit Before Tax (EBT)",       "profit_before_tax",             "subtotal"),
    ("Corporate Income Tax",          "corporate_income_tax",          "item"),
    ("Net Income",                    "net_income",                    "total"),
    ("  of which: Parent Company",    "net_income_parent",             "item"),
    ("  of which: Minority Interest", "net_income_minority",           "item"),
]

BS_ROWS = [
    ("ASSETS",                          None,                            "header"),
    ("Current Assets",                  None,                            "header"),
    ("Cash & Cash Equivalents",         "cash_and_equivalents",          "item"),
    ("Short-term Investments",          "short_term_investments",        "item"),
    ("Accounts Receivable",             "accounts_receivable",           "item"),
    ("Inventory",                       "inventory",                     "item"),
    ("Other Current Assets",            "other_current_assets",          "item"),
    ("Total Current Assets",            "total_current_assets",          "subtotal"),
    ("Non-Current Assets",              None,                            "header"),
    ("Long-term Receivables",           "long_term_receivables",         "item"),
    ("Fixed Assets (Net)",              "fixed_assets_net",              "item"),
    ("Investment Properties",           "investment_properties",         "item"),
    ("Long-term Investments",           "long_term_investments",         "item"),
    ("Other Non-Current Assets",        "other_non_current_assets",      "item"),
    ("Total Non-Current Assets",        "total_non_current_assets",      "subtotal"),
    ("TOTAL ASSETS",                    "total_assets",                  "total"),
    ("LIABILITIES & EQUITY",            None,                            "header"),
    ("Current Liabilities",             None,                            "header"),
    ("Short-term Loans",                "short_term_loans",              "item"),
    ("Accounts Payable",                "accounts_payable",              "item"),
    ("Other Current Liabilities",       "other_current_liabilities",     "item"),
    ("Total Current Liabilities",       "total_current_liabilities",     "subtotal"),
    ("Non-Current Liabilities",         None,                            "header"),
    ("Long-term Loans",                 "long_term_loans",               "item"),
    ("Other Non-Current Liabilities",   "other_non_current_liabilities", "item"),
    ("Total Non-Current Liabilities",   "total_non_current_liabilities", "subtotal"),
    ("Total Liabilities",               "total_liabilities",             "subtotal"),
    ("Equity",                          None,                            "header"),
    ("Charter Capital",                 "charter_capital",               "item"),
    ("Retained Earnings",               "retained_earnings",             "item"),
    ("Other Equity",                    "other_equity",                  "item"),
    ("Shareholders' Equity (Parent)",   "shareholders_equity_parent",    "subtotal"),
    ("Minority Interest",               "minority_interest",             "item"),
    ("Total Equity",                    "total_equity",                  "subtotal"),
    ("TOTAL LIABILITIES & EQUITY",      "total_liabilities_and_equity",  "total"),
]

CF_ROWS = [
    ("Operating Activities",            None,                            "header"),
    ("Net Income Before Tax",           "net_income_before_tax",         "item"),
    ("Depreciation & Amortization",     "depreciation_amortization",     "item"),
    ("Working Capital Changes",         "working_capital_changes",       "item"),
    ("Other Operating Adjustments",     "other_operating_adjustments",   "item"),
    ("Net Operating Cash Flow",         "net_operating_cf",              "subtotal"),
    ("Investing Activities",            None,                            "header"),
    ("Capital Expenditure (Capex)",     "capex",                         "item"),
    ("Proceeds from Asset Sales",       "proceeds_asset_sales",          "item"),
    ("Investment Outflows",             "investment_outflows",           "item"),
    ("Investment Inflows",              "investment_inflows",            "item"),
    ("Net Investing Cash Flow",         "net_investing_cf",              "subtotal"),
    ("Financing Activities",            None,                            "header"),
    ("Debt Raised",                     "debt_raised",                   "item"),
    ("Debt Repaid",                     "debt_repaid",                   "item"),
    ("Dividends Paid",                  "dividends_paid",                "item"),
    ("Other Financing",                 "other_financing",               "item"),
    ("Net Financing Cash Flow",         "net_financing_cf",              "subtotal"),
    ("Net Change in Cash",              "net_change_in_cash",            "subtotal"),
    ("Opening Cash Balance",            "opening_cash_balance",          "item"),
    ("Closing Cash Balance",            "closing_cash_balance",          "total"),
]

SHEET_ROWS = {
    "income_statement": IS_ROWS,
    "balance_sheet":    BS_ROWS,
    "cash_flow":        CF_ROWS,
}

FIELD_LABELS: dict[str, str] = {
    key: label
    for rows in (IS_ROWS, BS_ROWS, CF_ROWS)
    for label, key, _ in rows
    if key is not None
}


# ── Style primitives ──────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _side(style: str, color: str = C_BORDER) -> Side:
    return Side(style=style, color=color)

def _border(top=None, bottom=None, left=None, right=None) -> Border:
    return Border(
        top=top or Side(style=None),
        bottom=bottom or Side(style=None),
        left=left or Side(style=None),
        right=right or Side(style=None),
    )

def _top_thin()    -> Border: return _border(top=_side("thin"))
def _top_medium()  -> Border: return _border(top=_side("medium"))
def _top_bottom_medium() -> Border:
    return _border(top=_side("medium"), bottom=_side("medium"))
def _bottom_thin() -> Border: return _border(bottom=_side("thin"))
def _no_border()   -> Border: return Border()


def _setup_sheet(ws):
    """Apply display and print settings to a worksheet."""
    ws.sheet_view.showGridLines = False
    ws.print_options.gridLines  = False
    ws.print_options.headings   = False
    ws.page_setup.orientation   = "landscape"
    ws.page_setup.fitToPage     = True
    ws.page_setup.fitToWidth    = 1
    ws.page_setup.fitToHeight   = 0


# ── Formatted sheet builders ──────────────────────────────────────────────────

def _banner(ws, title: str, n_cols: int):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws["A1"]
    c.value     = title
    c.font      = Font(name="Calibri", bold=True, size=12, color=C_WHITE)
    c.fill      = _fill(C_BANNER)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26


def _col_headers(ws, years: list[int], first_col: str = "Line Item"):
    headers = [first_col] + [str(y) for y in years]
    for col, val in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=val)
        c.font      = Font(name="Calibri", bold=True, size=10, color=C_WHITE)
        c.fill      = _fill(C_COL_HDR)
        c.alignment = Alignment(
            horizontal="right" if col > 1 else "left",
            vertical="center",
            indent=1 if col == 1 else 0,
        )
        c.border = _border(bottom=_side("thin", color="4A6FA5"))
    ws.row_dimensions[2].height = 19


def _write_formatted_rows(ws, rows_def: list, data_by_year: dict, years: list[int]):
    """Write all canonical rows onto a formatted sheet."""

    # Pre-compute flat field lookup per year (excludes raw_lines)
    flat: dict[int, dict] = {}
    for y in years:
        d = data_by_year.get(y, {})
        flat[y] = {
            k: v
            for section in ("income_statement", "balance_sheet", "cash_flow")
            for k, v in d.get(section, {}).items()
            if k != "raw_lines"
        }

    for row_num, (label, field_key, row_type) in enumerate(rows_def, start=3):

        # ── Label cell (column A) ─────────────────────────────────────────
        lc = ws.cell(row=row_num, column=1, value=label)

        if row_type == "header":
            lc.font      = Font(name="Calibri", bold=True, size=10, color=C_TEXT_NAVY)
            lc.fill      = _fill(C_SECTION)
            lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            lc.border    = _bottom_thin()
            for col in range(2, len(years) + 2):
                nc        = ws.cell(row=row_num, column=col)
                nc.fill   = _fill(C_SECTION)
                nc.border = _bottom_thin()
            ws.row_dimensions[row_num].height = 14
            continue   # no data cells for section headers

        elif row_type == "total":
            lc.font      = Font(name="Calibri", bold=True, size=10, color=C_TEXT_NAVY)
            lc.fill      = _fill(C_TOTAL)
            lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            lc.border    = _top_bottom_medium()
            ws.row_dimensions[row_num].height = 17

        elif row_type == "subtotal":
            lc.font      = Font(name="Calibri", bold=True, size=10)
            lc.fill      = _fill(C_SUBTOTAL)
            lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            lc.border    = _top_thin()
            ws.row_dimensions[row_num].height = 16

        else:  # item
            indent       = 2 if label.startswith("  ") else 1
            lc.font      = Font(name="Calibri", size=10)
            lc.fill      = _fill(C_WHITE)
            lc.alignment = Alignment(horizontal="left", vertical="center", indent=indent)
            lc.border    = _no_border()
            ws.row_dimensions[row_num].height = 15

        # ── Value cells (year columns) ────────────────────────────────────
        for col_idx, year in enumerate(years, start=2):
            c = ws.cell(row=row_num, column=col_idx)
            c.alignment = Alignment(horizontal="right", vertical="center")

            if row_type == "total":
                c.fill   = _fill(C_TOTAL)
                c.font   = Font(name="Calibri", bold=True, size=10)
                c.border = _top_bottom_medium()
            elif row_type == "subtotal":
                c.fill   = _fill(C_SUBTOTAL)
                c.font   = Font(name="Calibri", bold=True, size=10)
                c.border = _top_thin()
            else:
                c.fill   = _fill(C_WHITE)
                c.font   = Font(name="Calibri", size=10)
                c.border = _no_border()

            value = flat[year].get(field_key)
            if value is None:
                c.value = "—"
                c.font  = Font(name="Calibri", size=10, color=C_TEXT_GRAY, italic=True,
                               bold=(row_type in ("subtotal", "total")))
            else:
                c.value         = value
                c.number_format = FIN_NUMBER_FORMAT


def _col_widths(ws, years: list[int], a_width: int = 40):
    ws.column_dimensions["A"].width = a_width
    for i in range(2, len(years) + 2):
        ws.column_dimensions[get_column_letter(i)].width = 16


# ── Raw sheet ─────────────────────────────────────────────────────────────────

def _build_raw_master(section_key: str, data_by_year: dict, years: list[int]) -> list[dict]:
    """Merge raw_lines across all years into a single ordered master list."""
    raw_by_year: dict[int, list[dict]] = {}
    for year in years:
        lines = data_by_year.get(year, {}).get(section_key, {}).get("raw_lines")
        if isinstance(lines, list) and lines:
            raw_by_year[year] = lines

    if not raw_by_year:
        return []

    master_year  = max(raw_by_year, key=lambda y: len(raw_by_year[y]))
    master_lines = raw_by_year[master_year]

    year_lookups: dict[int, dict] = {}
    for year, lines in raw_by_year.items():
        by_code:  dict[str, float | None] = {}
        by_label: dict[str, float | None] = {}
        for line in lines:
            code  = line.get("code")
            label = line.get("label", "")
            val   = line.get("value")
            if code:
                by_code[str(code)] = val
            by_label[label] = val
        year_lookups[year] = {"by_code": by_code, "by_label": by_label}

    result = []
    for line in master_lines:
        code  = str(line.get("code") or "").strip() or None
        label = line.get("label", "")
        ltype = line.get("type", "item")
        values: dict[int, float | None] = {}
        for year in years:
            if year not in year_lookups:
                values[year] = None
                continue
            lu  = year_lookups[year]
            val = None
            if code and code in lu["by_code"]:
                val = lu["by_code"][code]
            elif label in lu["by_label"]:
                val = lu["by_label"][label]
            values[year] = val
        result.append({"label": label, "code": code, "type": ltype, "values": values})

    # Append lines only present in non-master years
    master_codes  = {str(l.get("code") or "").strip() for l in master_lines if l.get("code")}
    master_labels = {l.get("label", "") for l in master_lines}
    for year, lines in raw_by_year.items():
        if year == master_year:
            continue
        for line in lines:
            code  = str(line.get("code") or "").strip() or None
            label = line.get("label", "")
            if (code and code in master_codes) or label in master_labels:
                continue
            values = {y: None for y in years}
            values[year] = line.get("value")
            result.append({"label": label, "code": code,
                           "type": line.get("type", "item"), "values": values})
            if code:
                master_codes.add(code)
            master_labels.add(label)

    return result


def _write_raw_sheet(ws, section_key: str, statement_name: str,
                     company_name: str, ticker: str,
                     data_by_year: dict, years: list[int]):
    """Write a raw extraction sheet showing every Vietnamese line item Claude extracted."""

    n_cols = 2 + len(years)   # Code | Vietnamese label | year...

    # ── Banner ────────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws["A1"]
    c.value     = f"Raw  ·  {statement_name}  ·  {company_name} ({ticker})  ·  Unit: Million VND"
    c.font      = Font(name="Calibri", bold=True, size=11, color=C_WHITE)
    c.fill      = _fill(C_RAW_HDR)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 24

    # ── Column headers ────────────────────────────────────────────────────────
    for col, val in enumerate(["Code", "Line Item (Vietnamese)"] + [str(y) for y in years], start=1):
        c = ws.cell(row=2, column=col, value=val)
        c.font      = Font(name="Calibri", bold=True, size=10, color=C_WHITE)
        c.fill      = _fill(C_RAW_HDR)
        c.alignment = Alignment(
            horizontal="center" if col == 1 else ("left" if col == 2 else "right"),
            vertical="center",
            indent=1 if col == 2 else 0,
        )
        c.border = _border(bottom=_side("thin", color="6B7B8D"))
    ws.row_dimensions[2].height = 19

    master = _build_raw_master(section_key, data_by_year, years)

    if not master:
        ws.cell(row=3, column=1,
                value="No raw_lines data available — run normalizer to generate extraction data.")
        _setup_sheet(ws)
        ws.freeze_panes = "C3"
        ws.sheet_properties.tabColor = C_RAW_HDR
        return

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_offset, entry in enumerate(master):
        row_num = row_offset + 3
        ltype   = entry["type"]
        label   = entry["label"]
        code    = entry["code"] or ""

        is_total    = ltype == "total"
        is_subtotal = ltype == "subtotal"
        ws.row_dimensions[row_num].height = 15

        if is_total:
            row_fill = _fill(C_TOTAL)
            font_kw  = dict(bold=True, size=10)
            brd      = _top_bottom_medium()
        elif is_subtotal:
            row_fill = _fill(C_SUBTOTAL)
            font_kw  = dict(bold=True, size=10)
            brd      = _top_thin()
        else:
            row_fill = _fill(C_WHITE)
            font_kw  = dict(size=10)
            brd      = _no_border()

        # Code cell
        cc = ws.cell(row=row_num, column=1, value=code)
        cc.font      = Font(name="Calibri", color="6B7B8D", **font_kw)
        cc.fill      = row_fill
        cc.alignment = Alignment(horizontal="center", vertical="center")
        cc.border    = brd

        # Vietnamese label cell
        lc = ws.cell(row=row_num, column=2, value=label)
        lc.font      = Font(name="Calibri", **font_kw)
        lc.fill      = row_fill
        lc.alignment = Alignment(
            horizontal="left", vertical="center",
            indent=0 if (is_total or is_subtotal) else 1,
        )
        lc.border = brd

        # Year value cells
        for col_idx, year in enumerate(years, start=3):
            value = entry["values"].get(year)
            c = ws.cell(row=row_num, column=col_idx)
            c.fill      = row_fill
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.border    = brd
            if value is None:
                c.value = "—"
                c.font  = Font(name="Calibri", color=C_TEXT_GRAY, italic=True, **font_kw)
            else:
                c.value         = value
                c.number_format = FIN_NUMBER_FORMAT
                c.font          = Font(name="Calibri", **font_kw)

    # ── Column widths + view ──────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 7    # Code — narrow
    ws.column_dimensions["B"].width = 56   # Vietnamese label — wide
    for i in range(3, 3 + len(years)):
        ws.column_dimensions[get_column_letter(i)].width = 17
    ws.freeze_panes = "C3"
    ws.sheet_properties.tabColor = C_RAW_HDR
    _setup_sheet(ws)


# ── English raw sheet ─────────────────────────────────────────────────────────

def _write_raw_en_sheet(ws, section_key: str, statement_name: str,
                        company_name: str, ticker: str,
                        data_by_year: dict, years: list[int]):
    """
    Write an English-translated raw sheet showing every line item Claude
    extracted, with:
      Col A  — Mã số (code)
      Col B  — English label (translated from Vietnamese)
      Col C  — Vietnamese label (original, for cross-reference)
      Col D+ — Year values
    Items that could not be translated are prefixed with ⚑.
    """
    n_cols = 3 + len(years)   # Code | EN label | VI label | year...

    # ── Banner ────────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws["A1"]
    c.value     = f"English  ·  {statement_name}  ·  {company_name} ({ticker})  ·  Unit: Million VND"
    c.font      = Font(name="Calibri", bold=True, size=11, color=C_WHITE)
    c.fill      = _fill(C_EN_HDR)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 24

    # ── Column headers ─────────────────────────────────────────────────────────
    col_headers = (
        ["Code", "Line Item (English)", "Line Item (Vietnamese)"]
        + [str(y) for y in years]
    )
    for col, val in enumerate(col_headers, start=1):
        c = ws.cell(row=2, column=col, value=val)
        c.font      = Font(name="Calibri", bold=True, size=10, color=C_WHITE)
        c.fill      = _fill(C_EN_HDR)
        c.alignment = Alignment(
            horizontal="center" if col == 1 else ("right" if col > 3 else "left"),
            vertical="center",
            indent=1 if col in (2, 3) else 0,
        )
        c.border = _border(bottom=_side("thin", color="52936E"))
    ws.row_dimensions[2].height = 19

    master = _build_raw_master(section_key, data_by_year, years)

    if not master:
        ws.cell(row=3, column=1,
                value="No raw_lines data available — run normalizer to generate extraction data.")
        _setup_sheet(ws)
        ws.freeze_panes = "D3"
        ws.sheet_properties.tabColor = C_EN_HDR
        return

    # ── Data rows ──────────────────────────────────────────────────────────────
    for row_offset, entry in enumerate(master):
        row_num = row_offset + 3
        ltype   = entry["type"]
        vi_label = entry["label"]
        en_label = _translate_vi(vi_label)
        code     = entry["code"] or ""

        is_total    = ltype == "total"
        is_subtotal = ltype == "subtotal"
        ws.row_dimensions[row_num].height = 15

        if is_total:
            row_fill = _fill(C_TOTAL)
            font_kw  = dict(bold=True, size=10)
            brd      = _top_bottom_medium()
        elif is_subtotal:
            row_fill = _fill(C_SUBTOTAL)
            font_kw  = dict(bold=True, size=10)
            brd      = _top_thin()
        else:
            row_fill = _fill(C_WHITE)
            font_kw  = dict(size=10)
            brd      = _no_border()

        is_untranslated = en_label.startswith("⚑")

        # Code cell
        cc = ws.cell(row=row_num, column=1, value=code)
        cc.font      = Font(name="Calibri", color="6B7B8D", **font_kw)
        cc.fill      = row_fill
        cc.alignment = Alignment(horizontal="center", vertical="center")
        cc.border    = brd

        # English label cell
        ec = ws.cell(row=row_num, column=2, value=en_label)
        ec.fill      = row_fill
        ec.alignment = Alignment(
            horizontal="left", vertical="center",
            indent=0 if (is_total or is_subtotal) else 1,
        )
        ec.border = brd
        if is_untranslated:
            ec.font = Font(name="Calibri", color="E07B39", **font_kw)   # amber — needs review
        else:
            ec.font = Font(name="Calibri", **font_kw)

        # Vietnamese label cell (gray, smaller — reference only)
        vc = ws.cell(row=row_num, column=3, value=vi_label)
        vc.font      = Font(name="Calibri", size=9, color="8A9BB0",
                            bold=font_kw.get("bold", False))
        vc.fill      = row_fill
        vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        vc.border    = brd

        # Year value cells
        for col_idx, year in enumerate(years, start=4):
            value = entry["values"].get(year)
            c = ws.cell(row=row_num, column=col_idx)
            c.fill      = row_fill
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.border    = brd
            if value is None:
                c.value = "—"
                c.font  = Font(name="Calibri", color=C_TEXT_GRAY, italic=True, **font_kw)
            else:
                c.value         = value
                c.number_format = FIN_NUMBER_FORMAT
                c.font          = Font(name="Calibri", **font_kw)

    # ── Column widths + view ───────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 7     # Code — narrow
    ws.column_dimensions["B"].width = 46    # English label
    ws.column_dimensions["C"].width = 46    # Vietnamese label
    for i in range(4, 4 + len(years)):
        ws.column_dimensions[get_column_letter(i)].width = 17
    ws.freeze_panes = "D3"
    ws.sheet_properties.tabColor = C_EN_HDR
    _setup_sheet(ws)


# ── Main export ───────────────────────────────────────────────────────────────

def export_company(ticker: str, company_data: list[dict],
                   years: list[int] | None = None) -> Path:
    """Build and save one financial-grade Excel workbook for a single company."""

    # Auto-detect years from available data if not supplied
    if years is None:
        years = sorted({entry["year"] for entry in company_data if entry.get("year")})
    if not years:
        years = DEFAULT_YEARS

    info  = get_company_info(ticker)
    name  = info["name"]
    color = info["color"]

    data_by_year: dict[int, dict] = {
        entry["year"]: entry
        for entry in company_data
        if entry.get("year")
    }

    wb = Workbook()
    wb.remove(wb.active)

    for section_key, sheet_name in SHEET_NAMES.items():
        n_cols = 1 + len(years)

        # ── Formatted sheet (canonical English, key metrics only) ─────────
        ws = wb.create_sheet(title=sheet_name)
        _setup_sheet(ws)
        _banner(ws,
                f"{sheet_name}  ·  {name} ({ticker})  ·  In Millions VND",
                n_cols)
        _col_headers(ws, years)
        _write_formatted_rows(ws, SHEET_ROWS[section_key], data_by_year, years)
        _col_widths(ws, years)
        ws.freeze_panes = "B3"
        ws.sheet_properties.tabColor = color

        # ── English raw sheet (all line items, translated) ─────────────────
        ws_en = wb.create_sheet(title=EN_RAW_SHEET_NAMES[section_key])
        _write_raw_en_sheet(ws_en, section_key, sheet_name,
                            name, ticker, data_by_year, years)

        # ── Vietnamese raw sheet (all line items, original labels) ─────────
        ws_raw = wb.create_sheet(title=RAW_SHEET_NAMES[section_key])
        _write_raw_sheet(ws_raw, section_key, sheet_name,
                         name, ticker, data_by_year, years)

    out_path = OUTPUT_DIR / f"{ticker}.xlsx"
    wb.save(out_path)
    print(f"  ✓ {out_path.name}  [{len(years)} year(s), 9 sheets]")
    return out_path


def run(all_data: list[dict] | None = None,
        years: list[int] | None = None) -> list[Path]:
    """Export all companies from normalized JSON data."""
    import json
    from config import DATA_PROCESSED

    if all_data is None:
        all_data = []
        for json_path in sorted(DATA_PROCESSED.glob("*.json")):
            try:
                all_data.append(json.loads(json_path.read_text(encoding="utf-8")))
            except Exception as e:
                print(f"  ⚠ Could not load {json_path.name}: {e}")

    if not all_data:
        print("No normalized data found. Run normalizer first.")
        return []

    by_ticker: dict[str, list[dict]] = {}
    for entry in all_data:
        t = entry.get("company", "UNKNOWN")
        by_ticker.setdefault(t, []).append(entry)

    saved = []
    for t in sorted(by_ticker):
        print(f"\n[{t}] Exporting...")
        saved.append(export_company(t, by_ticker[t], years=years))

    print(f"\n── Export complete: {len(saved)} workbook(s) ──")
    return saved


if __name__ == "__main__":
    run()
