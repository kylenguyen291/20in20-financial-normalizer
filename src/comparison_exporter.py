"""
src/comparison_exporter.py
Multi-Company Financial Comparison Workbook
============================================
Produces output/comparison.xlsx with 7 sheets:

  Overview       — Key ratios, latest year, all companies as columns
  Profitability  — Margins, ROE, ROA, EBITDA margin  (all years)
  Leverage       — Debt ratios, interest coverage     (all years)
  Liquidity      — Current & quick ratios             (all years)
  Efficiency     — Turnover & days metrics            (all years)
  Cash Flow      — CF quality, FCF margin             (all years)
  Key Financials — Absolute figures in Million VND    (all years)

Usage:
    from src.comparison_exporter import build_comparison
    build_comparison(all_data, years=[2022, 2023, 2024])
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from config import get_company_info, OUTPUT_DIR

# ── Color palette ──────────────────────────────────────────────────────────────
C_BANNER    = "1B4F72"   # Deep navy-blue  — comparison workbook banner
C_CO_HDR    = None       # Company header uses each company's own color
C_YEAR_HDR  = "34495E"   # Dark slate      — year sub-header row
C_CAT_HDR   = "EEF2F8"   # Ghost white     — category section dividers
C_ALT       = "F5F8FD"   # Very light blue — alternating metric rows
C_WHITE     = "FFFFFF"
C_NAVY      = "1F3864"
C_GRAY      = "9CA3AF"
C_BORDER    = "C5D3E0"

# ── Number formats ─────────────────────────────────────────────────────────────
FMT_PCT  = '0.0%;(0.0%);"-"'
FMT_X    = '0.0"x";(0.0"x");"-"'
FMT_DAYS = '0.0"d";(0.0"d");"-"'
FMT_ABS  = '#,##0;(#,##0);"-"'

# ── Style helpers ──────────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _side(style, color=C_BORDER):
    return Side(style=style, color=color)

def _border(top=None, bottom=None):
    return Border(
        top=top    or Side(style=None),
        bottom=bottom or Side(style=None),
        left=Side(style=None),
        right=Side(style=None),
    )

def _setup(ws):
    ws.sheet_view.showGridLines = False
    ws.print_options.gridLines  = False
    ws.page_setup.orientation   = "landscape"
    ws.page_setup.fitToPage     = True
    ws.page_setup.fitToWidth    = 1
    ws.page_setup.fitToHeight   = 0


# ── Math helpers ───────────────────────────────────────────────────────────────

def _div(a, b):
    """Safe division — returns None on zero denominator or missing values."""
    if a is None or b is None or b == 0:
        return None
    return a / b

def _sum2(a, b):
    """Safe addition of two nullable values."""
    if a is None and b is None:
        return None
    return (a or 0) + (b or 0)


# ── Ratio computation ──────────────────────────────────────────────────────────

def _compute(entry: dict) -> dict:
    """
    Compute all ratios and derived metrics from one company-year data entry.
    All inputs are in Million VND. Returns a flat dict of computed values.
    """
    def _v(section, field):
        return entry.get(section, {}).get(field)

    # Raw inputs
    rev   = _v("income_statement", "net_revenue")
    gp    = _v("income_statement", "gross_profit")
    ebit  = _v("income_statement", "operating_profit")
    ni    = _v("income_statement", "net_income")
    cogs  = _v("income_statement", "cost_of_goods_sold")
    int_e = _v("income_statement", "interest_expense")

    ta    = _v("balance_sheet", "total_assets")
    te    = _v("balance_sheet", "total_equity")
    tca   = _v("balance_sheet", "total_current_assets")
    tcl   = _v("balance_sheet", "total_current_liabilities")
    inv   = _v("balance_sheet", "inventory")
    ar    = _v("balance_sheet", "accounts_receivable")
    ap    = _v("balance_sheet", "accounts_payable")
    cash  = _v("balance_sheet", "cash_and_equivalents")
    stl   = _v("balance_sheet", "short_term_loans")
    ltl   = _v("balance_sheet", "long_term_loans")
    tl    = _v("balance_sheet", "total_liabilities")

    dep   = _v("cash_flow", "depreciation_amortization")
    opcf  = _v("cash_flow", "net_operating_cf")
    capex = _v("cash_flow", "capex")

    # Derived intermediates
    ebitda = _sum2(ebit, dep)
    td     = _sum2(stl, ltl)
    nd     = _sum2(td, -cash if cash is not None else None)
    fcf    = _sum2(opcf, capex)
    tca_ex_inv = _sum2(tca, -inv if inv is not None else None)

    return {
        # ── Profitability ──────────────────────────────────────────────────────
        "Gross Margin":    (_div(gp,    rev),    FMT_PCT),
        "EBIT Margin":     (_div(ebit,  rev),    FMT_PCT),
        "Net Margin":      (_div(ni,    rev),    FMT_PCT),
        "ROE":             (_div(ni,    te),     FMT_PCT),
        "ROA":             (_div(ni,    ta),     FMT_PCT),
        "EBITDA":          (ebitda,              FMT_ABS),
        "EBITDA Margin":   (_div(ebitda, rev),   FMT_PCT),

        # ── Leverage ───────────────────────────────────────────────────────────
        "Total Debt":      (td,                  FMT_ABS),
        "Net Debt":        (nd,                  FMT_ABS),
        "Debt / Equity":   (_div(td,    te),     FMT_X),
        "Net Debt / EBITDA": (_div(nd,  ebitda), FMT_X),
        "Interest Coverage": (_div(ebit, int_e), FMT_X),
        "Liabilities / Equity": (_div(tl, te),  FMT_X),

        # ── Liquidity ──────────────────────────────────────────────────────────
        "Current Ratio":   (_div(tca, tcl),           FMT_X),
        "Quick Ratio":     (_div(tca_ex_inv, tcl),    FMT_X),
        "Cash Ratio":      (_div(cash, tcl),           FMT_X),

        # ── Efficiency ─────────────────────────────────────────────────────────
        "Asset Turnover":  (_div(rev, ta),             FMT_X),
        "Inventory Days":  (_div(inv,  _div(cogs, 365)), FMT_DAYS),
        "Receivable Days": (_div(ar,   _div(rev,  365)), FMT_DAYS),
        "Payable Days":    (_div(ap,   _div(cogs, 365)), FMT_DAYS),

        # ── Cash Flow ──────────────────────────────────────────────────────────
        "Operating CF":    (opcf,                      FMT_ABS),
        "Free Cash Flow":  (fcf,                       FMT_ABS),
        "CF / Net Income": (_div(opcf, ni),            FMT_X),
        "Capex":           (capex,                     FMT_ABS),
        "Capex / Revenue": (_div(abs(capex) if capex else None, rev), FMT_PCT),
        "FCF Margin":      (_div(fcf, rev),            FMT_PCT),

        # ── Key Financials (absolute) ──────────────────────────────────────────
        "Revenue":         (rev,   FMT_ABS),
        "Gross Profit":    (gp,    FMT_ABS),
        "Net Income":      (ni,    FMT_ABS),
        "Total Assets":    (ta,    FMT_ABS),
        "Total Equity":    (te,    FMT_ABS),
        "Total Liabilities": (tl,  FMT_ABS),
    }


# ── Sheet definitions ──────────────────────────────────────────────────────────
# Each sheet: list of (category_header | metric_name, key_in_compute_dict)
# category_header rows have key=None (no data, just a divider)

SHEETS = {
    "Profitability": [
        ("Margin Ratios",     None),
        ("Gross Margin",      "Gross Margin"),
        ("EBIT Margin",       "EBIT Margin"),
        ("Net Margin",        "Net Margin"),
        ("EBITDA Margin",     "EBITDA Margin"),
        ("Return Ratios",     None),
        ("ROE",               "ROE"),
        ("ROA",               "ROA"),
        ("EBITDA (Mln VND)",  "EBITDA"),
    ],
    "Leverage": [
        ("Debt Ratios",           None),
        ("Total Debt (Mln VND)",  "Total Debt"),
        ("Net Debt (Mln VND)",    "Net Debt"),
        ("Debt / Equity",         "Debt / Equity"),
        ("Net Debt / EBITDA",     "Net Debt / EBITDA"),
        ("Coverage",              None),
        ("Interest Coverage",     "Interest Coverage"),
        ("Liabilities / Equity",  "Liabilities / Equity"),
    ],
    "Liquidity": [
        ("Liquidity Ratios",  None),
        ("Current Ratio",     "Current Ratio"),
        ("Quick Ratio",       "Quick Ratio"),
        ("Cash Ratio",        "Cash Ratio"),
    ],
    "Efficiency": [
        ("Turnover",          None),
        ("Asset Turnover",    "Asset Turnover"),
        ("Working Capital Days", None),
        ("Inventory Days",    "Inventory Days"),
        ("Receivable Days",   "Receivable Days"),
        ("Payable Days",      "Payable Days"),
    ],
    "Cash Flow": [
        ("Cash Generation",       None),
        ("Operating CF (Mln VND)","Operating CF"),
        ("Free Cash Flow (Mln VND)","Free Cash Flow"),
        ("CF / Net Income",       "CF / Net Income"),
        ("FCF Margin",            "FCF Margin"),
        ("Investment",            None),
        ("Capex (Mln VND)",       "Capex"),
        ("Capex / Revenue",       "Capex / Revenue"),
    ],
    "Key Financials": [
        ("Income Statement",      None),
        ("Revenue (Mln VND)",     "Revenue"),
        ("Gross Profit (Mln VND)","Gross Profit"),
        ("EBITDA (Mln VND)",      "EBITDA"),
        ("Net Income (Mln VND)",  "Net Income"),
        ("Balance Sheet",         None),
        ("Total Assets (Mln VND)","Total Assets"),
        ("Total Equity (Mln VND)","Total Equity"),
        ("Total Liabilities (Mln VND)", "Total Liabilities"),
        ("Total Debt (Mln VND)",  "Total Debt"),
        ("Net Debt (Mln VND)",    "Net Debt"),
        ("Cash Flow",             None),
        ("Operating CF (Mln VND)","Operating CF"),
        ("Free Cash Flow (Mln VND)","Free Cash Flow"),
    ],
}

OVERVIEW_METRICS = [
    ("Profitability",     None),
    ("Gross Margin",      "Gross Margin"),
    ("EBIT Margin",       "EBIT Margin"),
    ("Net Margin",        "Net Margin"),
    ("ROE",               "ROE"),
    ("EBITDA Margin",     "EBITDA Margin"),
    ("Leverage",          None),
    ("Debt / Equity",     "Debt / Equity"),
    ("Net Debt / EBITDA", "Net Debt / EBITDA"),
    ("Interest Coverage", "Interest Coverage"),
    ("Liquidity",         None),
    ("Current Ratio",     "Current Ratio"),
    ("Quick Ratio",       "Quick Ratio"),
    ("Efficiency",        None),
    ("Asset Turnover",    "Asset Turnover"),
    ("Inventory Days",    "Inventory Days"),
    ("Receivable Days",   "Receivable Days"),
    ("Cash Flow",         None),
    ("FCF Margin",        "FCF Margin"),
    ("CF / Net Income",   "CF / Net Income"),
]


# ── Sheet writers ──────────────────────────────────────────────────────────────

def _banner(ws, text: str, n_cols: int):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws["A1"]
    c.value     = text
    c.font      = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    c.fill      = _fill(C_BANNER)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26


def _write_comparison_sheet(ws, sheet_rows: list, computed: dict,
                             tickers: list, years: list,
                             ticker_colors: dict, sheet_title: str):
    """
    Write a ratio comparison sheet.

    Layout:
      Row 1 : Banner
      Row 2 : Metric  |  ← Ticker A →  |  ← Ticker B →  |  ...
      Row 3 :          |  yr1 | yr2 ... |  yr1 | yr2 ... |
      Row 4+: data rows
    """
    n_yr   = len(years)
    n_co   = len(tickers)
    n_data = n_co * n_yr
    n_cols = 1 + n_data

    _setup(ws)
    _banner(ws, f"{sheet_title}  ·  Financial Comparison  ·  Unit: Million VND", n_cols)

    # ── Row 2: Company group headers ──────────────────────────────────────────
    ws.row_dimensions[2].height = 20
    # blank label cell
    lc = ws.cell(row=2, column=1, value="")
    lc.fill = _fill(C_YEAR_HDR)
    lc.border = _border(bottom=_side("thin", "4A6FA5"))

    for co_idx, ticker in enumerate(tickers):
        col_start = 2 + co_idx * n_yr
        col_end   = col_start + n_yr - 1
        color     = ticker_colors.get(ticker, "2E4057")

        if col_start == col_end:
            c = ws.cell(row=2, column=col_start, value=ticker)
        else:
            ws.merge_cells(start_row=2, start_column=col_start,
                           end_row=2, end_column=col_end)
            c = ws.cell(row=2, column=col_start, value=ticker)

        c.font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill      = _fill(color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _border(bottom=_side("thin"))

        for col in range(col_start + 1, col_end + 1):
            nc = ws.cell(row=2, column=col)
            nc.fill   = _fill(color)
            nc.border = _border(bottom=_side("thin"))

    # ── Row 3: Year sub-headers ───────────────────────────────────────────────
    ws.row_dimensions[3].height = 18
    lc = ws.cell(row=3, column=1, value="Metric")
    lc.font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    lc.fill      = _fill(C_YEAR_HDR)
    lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    lc.border    = _border(bottom=_side("thin", "4A6FA5"))

    for co_idx in range(n_co):
        for yr_idx, year in enumerate(years):
            col = 2 + co_idx * n_yr + yr_idx
            c   = ws.cell(row=3, column=col, value=str(year))
            c.font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            c.fill      = _fill(C_YEAR_HDR)
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.border    = _border(bottom=_side("thin", "4A6FA5"))

    # ── Data rows ─────────────────────────────────────────────────────────────
    alt = False
    for row_offset, (label, metric_key) in enumerate(sheet_rows):
        row = 4 + row_offset
        ws.row_dimensions[row].height = 15

        is_cat = metric_key is None

        if is_cat:
            # Category divider
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=n_cols)
            c = ws.cell(row=row, column=1, value=label)
            c.font      = Font(name="Calibri", bold=True, size=10, color=C_NAVY)
            c.fill      = _fill(C_CAT_HDR)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c.border    = _border(bottom=_side("thin", C_BORDER))
            ws.row_dimensions[row].height = 14
            alt = False
            continue

        row_fill = C_ALT if alt else C_WHITE
        alt = not alt

        # Label cell
        lc = ws.cell(row=row, column=1, value=label)
        lc.font      = Font(name="Calibri", size=10)
        lc.fill      = _fill(row_fill)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=2)

        # Value cells
        for co_idx, ticker in enumerate(tickers):
            for yr_idx, year in enumerate(years):
                col   = 2 + co_idx * n_yr + yr_idx
                entry = computed.get((ticker, year), {})
                cell_val, fmt = entry.get(metric_key, (None, FMT_ABS))

                c = ws.cell(row=row, column=col)
                c.fill      = _fill(row_fill)
                c.alignment = Alignment(horizontal="right", vertical="center")

                if cell_val is None:
                    c.value = "—"
                    c.font  = Font(name="Calibri", size=10, color=C_GRAY, italic=True)
                else:
                    c.value         = cell_val
                    c.number_format = fmt
                    c.font          = Font(name="Calibri", size=10)

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 26
    for i in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 12

    ws.freeze_panes = "B4"


def _write_overview(ws, computed: dict, tickers: list, years: list,
                    ticker_colors: dict):
    """
    Overview sheet: all years, all key ratios — same multi-year layout as other sheets.
    Color-scale conditional formatting highlights best/worst values per row.
    """
    _write_comparison_sheet(ws, OVERVIEW_METRICS, computed,
                            tickers, years, ticker_colors, "Overview")

    # Color scale across entire data area (skipping category divider rows)
    n_yr   = len(years)
    n_co   = len(tickers)
    n_cols = 1 + n_co * n_yr
    first_col = get_column_letter(2)
    last_col  = get_column_letter(n_cols)
    last_row  = 3 + len(OVERVIEW_METRICS)
    ws.conditional_formatting.add(
        f"{first_col}4:{last_col}{last_row}",
        ColorScaleRule(
            start_type="min",  start_color="FCE4D6",
            mid_type="num",    mid_value=0, mid_color="FFFFFF",
            end_type="max",    end_color="E2EFDA",
        ),
    )
    ws.sheet_properties.tabColor = C_BANNER


# ── Insights engine ───────────────────────────────────────────────────────────

def _trend(vals: list) -> str:
    """Return '▲ improving', '▼ declining', or '→ stable' for a series of values."""
    clean = [v for v in vals if v is not None]
    if len(clean) < 2:
        return "→ n/a"
    delta = clean[-1] - clean[0]
    pct   = abs(delta / clean[0]) if clean[0] != 0 else 0
    if pct < 0.03:
        return "→ stable"
    return "▲ improving" if delta > 0 else "▼ declining"


def _fmt_val(val, fmt):
    if val is None:
        return "n/a"
    if fmt == FMT_PCT:
        return f"{val*100:.1f}%"
    if fmt == FMT_X:
        return f"{val:.1f}x"
    if fmt == FMT_DAYS:
        return f"{val:.0f}d"
    return f"{val:,.0f}"


def _generate_insights(computed: dict, tickers: list, years: list) -> list[str]:
    """
    Analyse the computed ratios and return a list of plain-English insight strings.
    Each string is one bullet point for the Insights sheet.
    """
    insights = []
    sorted_years = sorted(years)

    def _vals(ticker, metric):
        return [computed.get((ticker, y), {}).get(metric, (None, None))[0]
                for y in sorted_years]

    def _latest(ticker, metric):
        for y in reversed(sorted_years):
            v = computed.get((ticker, y), {}).get(metric, (None, None))[0]
            if v is not None:
                return v, y
        return None, None

    # ── Profitability ──────────────────────────────────────────────────────────
    insights.append("PROFITABILITY")

    # Best net margin latest year
    margins = {t: _latest(t, "Net Margin") for t in tickers}
    valid   = {t: (v, y) for t, (v, y) in margins.items() if v is not None}
    if valid:
        best  = max(valid, key=lambda t: valid[t][0])
        worst = min(valid, key=lambda t: valid[t][0])
        insights.append(
            f"  {best} has the highest net margin at {_fmt_val(valid[best][0], FMT_PCT)} "
            f"({valid[best][1]}), versus {worst} at {_fmt_val(valid[worst][0], FMT_PCT)}."
        )

    # Margin trends
    for t in tickers:
        vals  = _vals(t, "Net Margin")
        trend = _trend(vals)
        v0, v1 = vals[0], vals[-1]
        if "improving" in trend or "declining" in trend:
            insights.append(
                f"  {t} net margin is {trend.split()[1]} "
                f"({_fmt_val(v0, FMT_PCT)} → {_fmt_val(v1, FMT_PCT)} over {sorted_years[0]}–{sorted_years[-1]})."
            )

    # ROE comparison
    roes = {t: _latest(t, "ROE") for t in tickers}
    valid_roe = {t: (v, y) for t, (v, y) in roes.items() if v is not None}
    if valid_roe:
        best_roe = max(valid_roe, key=lambda t: valid_roe[t][0])
        insights.append(
            f"  {best_roe} leads on ROE at {_fmt_val(valid_roe[best_roe][0], FMT_PCT)} "
            f"({valid_roe[best_roe][1]})."
        )

    # ── Leverage ───────────────────────────────────────────────────────────────
    insights.append("")
    insights.append("LEVERAGE & SOLVENCY")

    for t in tickers:
        nd_ebitda, yr = _latest(t, "Net Debt / EBITDA")
        de,         _  = _latest(t, "Debt / Equity")
        if nd_ebitda is not None:
            flag = " ⚠ High leverage" if nd_ebitda > 3.0 else (" ✓ Manageable" if nd_ebitda < 1.5 else "")
            insights.append(
                f"  {t}: Net Debt/EBITDA = {_fmt_val(nd_ebitda, FMT_X)}"
                + (f", D/E = {_fmt_val(de, FMT_X)}" if de is not None else "")
                + f".{flag}"
            )

    # Interest coverage warnings
    for t in tickers:
        cov, yr = _latest(t, "Interest Coverage")
        if cov is not None and cov < 3.0:
            insights.append(
                f"  ⚠  {t} interest coverage is thin at {_fmt_val(cov, FMT_X)} "
                f"({yr}) — monitor debt service capacity."
            )

    # Leverage trends
    for t in tickers:
        vals  = _vals(t, "Debt / Equity")
        trend = _trend(vals)
        if "improving" in trend or "declining" in trend:
            direction = "building up debt" if "declining" in trend else "deleveraging"
            insights.append(f"  {t} D/E is {trend.split()[1]} — {direction}.")

    # ── Liquidity ──────────────────────────────────────────────────────────────
    insights.append("")
    insights.append("LIQUIDITY")

    for t in tickers:
        cr, yr = _latest(t, "Current Ratio")
        qr, _  = _latest(t, "Quick Ratio")
        if cr is not None:
            flag = " ⚠ Below 1 — potential short-term stress" if cr < 1.0 else ""
            insights.append(
                f"  {t}: Current Ratio = {_fmt_val(cr, FMT_X)}"
                + (f", Quick Ratio = {_fmt_val(qr, FMT_X)}" if qr is not None else "")
                + f" ({yr}).{flag}"
            )

    # ── Efficiency ─────────────────────────────────────────────────────────────
    insights.append("")
    insights.append("EFFICIENCY")

    # Inventory days comparison
    inv_days = {t: _latest(t, "Inventory Days") for t in tickers}
    valid_inv = {t: (v, y) for t, (v, y) in inv_days.items() if v is not None}
    if len(valid_inv) > 1:
        slowest = max(valid_inv, key=lambda t: valid_inv[t][0])
        fastest = min(valid_inv, key=lambda t: valid_inv[t][0])
        insights.append(
            f"  {fastest} turns inventory fastest ({_fmt_val(valid_inv[fastest][0], FMT_DAYS)}), "
            f"while {slowest} holds stock for {_fmt_val(valid_inv[slowest][0], FMT_DAYS)} on average."
        )

    # Receivable days
    ar_days = {t: _latest(t, "Receivable Days") for t in tickers}
    valid_ar = {t: (v, y) for t, (v, y) in ar_days.items() if v is not None}
    if len(valid_ar) > 1:
        slowest = max(valid_ar, key=lambda t: valid_ar[t][0])
        fastest = min(valid_ar, key=lambda t: valid_ar[t][0])
        insights.append(
            f"  {fastest} collects receivables in {_fmt_val(valid_ar[fastest][0], FMT_DAYS)}, "
            f"versus {slowest} at {_fmt_val(valid_ar[slowest][0], FMT_DAYS)}."
        )

    # Asset turnover
    at_vals = {t: _latest(t, "Asset Turnover") for t in tickers}
    valid_at = {t: (v, y) for t, (v, y) in at_vals.items() if v is not None}
    if valid_at:
        best_at = max(valid_at, key=lambda t: valid_at[t][0])
        insights.append(
            f"  {best_at} uses assets most efficiently (turnover {_fmt_val(valid_at[best_at][0], FMT_X)})."
        )

    # ── Cash Flow ──────────────────────────────────────────────────────────────
    insights.append("")
    insights.append("CASH FLOW QUALITY")

    for t in tickers:
        cf_ni, yr = _latest(t, "CF / Net Income")
        fcf_m, _  = _latest(t, "FCF Margin")
        if cf_ni is not None:
            quality = "strong" if cf_ni > 1.0 else ("weak" if cf_ni < 0.5 else "moderate")
            insights.append(
                f"  {t}: CF/Net Income = {_fmt_val(cf_ni, FMT_X)} — {quality} earnings quality"
                + (f", FCF Margin = {_fmt_val(fcf_m, FMT_PCT)}" if fcf_m is not None else "")
                + f" ({yr})."
            )

    # FCF trends
    for t in tickers:
        vals  = _vals(t, "FCF Margin")
        trend = _trend(vals)
        if "improving" in trend or "declining" in trend:
            insights.append(f"  {t} FCF margin trend: {trend}.")

    # ── Summary ────────────────────────────────────────────────────────────────
    insights.append("")
    insights.append("SUMMARY")

    # Overall best performer by net margin
    all_latest = {t: _latest(t, "Net Margin")[0] for t in tickers}
    valid_all  = {t: v for t, v in all_latest.items() if v is not None}
    if valid_all:
        best_overall = max(valid_all, key=lambda t: valid_all[t])
        insights.append(
            f"  On profitability, {best_overall} is the strongest performer in this peer group."
        )

    return insights


def _write_insights(ws, insights: list[str], tickers: list, years: list):
    """Write the Insights sheet as formatted text observations."""
    _setup(ws)
    yr_range = f"{min(years)}–{max(years)}" if len(years) > 1 else str(years[0])
    tickers_str = ", ".join(tickers)

    # Banner
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = f"Insights  ·  {tickers_str}  ·  {yr_range}"
    c.font      = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    c.fill      = _fill(C_BANNER)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    # Sub-header
    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value     = "AI-generated observations from normalised financial data  ·  Review before use"
    c.font      = Font(name="Calibri", italic=True, size=9, color="6B7B8D")
    c.fill      = _fill("F0F4F8")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 16

    # Insight rows
    for row_offset, line in enumerate(insights):
        row = 3 + row_offset
        ws.row_dimensions[row].height = 16 if line else 8

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws.cell(row=row, column=1, value=line)

        if not line:
            c.value = None
            continue

        is_heading = not line.startswith(" ")

        if is_heading:
            c.font      = Font(name="Calibri", bold=True, size=10, color=C_NAVY)
            c.fill      = _fill(C_CAT_HDR)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c.border    = _border(bottom=_side("thin", C_BORDER))
            ws.row_dimensions[row].height = 18
        else:
            is_warning = "⚠" in line
            is_good    = "✓" in line or "▲" in line
            font_color = "C0392B" if is_warning else ("27AE60" if is_good else "2C3E50")
            row_fill   = "FEF9E7" if is_warning else ("EAFAF1" if is_good else C_WHITE)
            c.font      = Font(name="Calibri", size=10, color=font_color)
            c.fill      = _fill(row_fill)
            c.alignment = Alignment(horizontal="left", vertical="center",
                                    wrap_text=True, indent=1)

    ws.column_dimensions["A"].width = 90
    ws.sheet_properties.tabColor = "E74C3C"


# ── Entry point ────────────────────────────────────────────────────────────────

def build_comparison(all_data: list[dict],
                     years: list[int] | None = None) -> Path:
    """
    Build and save the multi-company comparison workbook.

    Parameters
    ----------
    all_data : list of normalized JSON records (one per company × year)
    years    : year list to include; auto-detected from data if None

    Returns
    -------
    Path to the saved workbook
    """
    if not all_data:
        print("  ✗ No data for comparison workbook.")
        return None

    # ── Resolve tickers and years ──────────────────────────────────────────────
    tickers = sorted({d["company"] for d in all_data if d.get("company")})
    if not years:
        years = sorted({d["year"] for d in all_data if d.get("year")})

    if not tickers or not years:
        print("  ✗ No tickers or years found in data.")
        return None

    # ── Gather per-company colors ──────────────────────────────────────────────
    ticker_colors = {t: get_company_info(t)["color"] for t in tickers}

    # ── Pre-compute all ratios ─────────────────────────────────────────────────
    computed: dict[tuple, dict] = {}
    for entry in all_data:
        t = entry.get("company")
        y = entry.get("year")
        if t and y:
            computed[(t, y)] = _compute(entry)

    # ── Build workbook ─────────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    # Insights — first tab, quick read
    ws_ins = wb.create_sheet("Insights")
    insights = _generate_insights(computed, tickers, years)
    _write_insights(ws_ins, insights, tickers, years)

    # Overview — all years, all key ratios
    ws_ov = wb.create_sheet("Overview")
    _write_overview(ws_ov, computed, tickers, years, ticker_colors)

    # Category sheets
    sheet_colors = {
        "Profitability": "1A5276",   # dark blue
        "Leverage":      "784212",   # dark brown
        "Liquidity":     "145A32",   # dark green
        "Efficiency":    "4A235A",   # dark purple
        "Cash Flow":     "1B4F72",   # navy
        "Key Financials":"2E4057",   # slate
    }

    for sheet_name, sheet_rows in SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = sheet_colors.get(sheet_name, C_BANNER)
        _write_comparison_sheet(ws, sheet_rows, computed,
                                tickers, years, ticker_colors, sheet_name)

    out_path = OUTPUT_DIR / "comparison.xlsx"
    wb.save(out_path)
    n_pairs = len(computed)
    print(f"  ✓ comparison.xlsx  [{len(tickers)} companies × {len(years)} years = {n_pairs} data points, 8 sheets]")
    return out_path
